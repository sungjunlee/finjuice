"""Tests for finjuice pipeline commands (ingest, tag, transfer, export, refresh, all)."""

import json
from pathlib import Path
from unittest.mock import patch

import polars as pl
import pytest
from typer.testing import CliRunner

from finjuice.pipeline.cli.main import app

runner = CliRunner()


def _snapshot_tree(root: Path) -> dict[str, int]:
    """Capture a filesystem snapshot for no-write dry-run assertions."""
    snapshot = {".": root.stat().st_mtime_ns}
    for path in sorted(root.rglob("*")):
        snapshot[str(path.relative_to(root))] = path.stat().st_mtime_ns
    return snapshot


def _read_audit_events(data_dir: Path) -> list[dict[str, object]]:
    """Read JSONL audit events from a test data directory."""
    audit_path = data_dir / ".execution_audit.jsonl"
    if not audit_path.exists():
        return []
    return [json.loads(line) for line in audit_path.read_text(encoding="utf-8").splitlines()]


@pytest.fixture
def data_dir_with_xlsx(tmp_path: Path) -> Path:
    """Create data directory with a sample XLSX file in imports/.

    Note: Banksalad exports have transaction data on the second sheet (index 1)
    with title "가계부 내역". The first sheet contains summary data.
    """
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    (data_dir / "transactions").mkdir()
    imports_dir = data_dir / "imports"
    imports_dir.mkdir()
    (data_dir / "exports").mkdir()
    # Create minimal rules.yaml to mark the data directory as initialized
    (data_dir / "rules.yaml").write_text("version: 1\nrules: []\n")

    # Create a sample XLSX file matching Banksalad export format
    import openpyxl

    wb = openpyxl.Workbook()
    # First sheet (summary, not used)
    ws_summary = wb.active
    ws_summary.title = "요약"
    ws_summary.cell(row=1, column=1, value="요약 데이터")

    # Second sheet - transaction data (this is what Banksalad uses)
    ws = wb.create_sheet("가계부 내역")

    # Header row
    headers = [
        "날짜",
        "시간",
        "타입",
        "대분류",
        "소분류",
        "내용",
        "금액",
        "화폐",
        "결제수단",
        "메모",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    # Data row
    ws.cell(row=2, column=1, value="2024-10-01")
    ws.cell(row=2, column=2, value="10:00")
    ws.cell(row=2, column=3, value="지출")
    ws.cell(row=2, column=4, value="식비")
    ws.cell(row=2, column=5, value="카페")
    ws.cell(row=2, column=6, value="스타벅스")
    ws.cell(row=2, column=7, value=5000)
    ws.cell(row=2, column=8, value="KRW")
    ws.cell(row=2, column=9, value="신한카드")
    ws.cell(row=2, column=10, value="")
    wb.save(imports_dir / "test_export.xlsx")

    return data_dir


@pytest.fixture
def data_dir_with_transactions(tmp_path: Path) -> Path:
    """Create data directory with existing transaction CSV files."""
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    (data_dir / "imports").mkdir()
    (data_dir / "exports").mkdir()

    # Create partition structure
    partition_dir = data_dir / "transactions" / "2024" / "10"
    partition_dir.mkdir(parents=True)

    # Create sample DataFrame with test data
    df = pl.DataFrame(
        {
            "row_hash": ["abc12345678901234", "def45678901234567", "ghi78901234567890"],
            "date": ["2024-10-01", "2024-10-02", "2024-10-03"],
            "time": ["10:00", "12:00", "14:00"],
            "type_raw": ["지출", "지출", "지출"],
            "type_norm": ["expense", "expense", "expense"],
            "major_raw": ["식비", "식비", "교통"],
            "minor_raw": ["카페", "음식점", "택시"],
            "merchant_raw": ["스타벅스", "맥도날드", "카카오택시"],
            "memo_raw": ["", "", ""],
            "amount": [-5000.0, -8000.0, -15000.0],
            "account": ["신한카드", "삼성카드", "현금"],
            "currency": ["KRW"] * 3,
            "counterparty": ["", "", ""],
            "datetime": [
                "2024-10-01T10:00:00",
                "2024-10-02T12:00:00",
                "2024-10-03T14:00:00",
            ],
            "tags_rule": ["[]"] * 3,
            "tags_ai": ["[]"] * 3,
            "tags_manual": ["[]"] * 3,
            "tags_final": ["[]"] * 3,
            "confidence": [0.0, 0.0, 0.0],
            "needs_review": [0, 0, 0],
            "is_transfer": [0, 0, 0],
            "transfer_group_id": ["", "", ""],
            "file_id": ["241001_1"] * 3,
            "source_row": [1, 2, 3],
        }
    )

    df.write_csv(partition_dir / "transactions.csv")

    # Create rules.yaml with sample rules
    rules_content = """version: 1
rules:
  - name: cafe_starbucks
    match: "스타벅스|STARBUCKS"
    fields: [merchant_raw]
    tags: ["카페", "커피"]
    priority: 80
"""
    (data_dir / "rules.yaml").write_text(rules_content)

    return data_dir


@pytest.fixture
def data_dir_empty(tmp_path: Path) -> Path:
    """Create an empty data directory structure."""
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    (data_dir / "transactions").mkdir()
    (data_dir / "imports").mkdir()
    (data_dir / "exports").mkdir()
    # Create minimal rules.yaml to mark the data directory as initialized
    (data_dir / "rules.yaml").write_text("version: 1\nrules: []\n")
    return data_dir


class TestIngestCommand:
    """Tests for the ingest command."""

    def test_ingest_no_xlsx_files(self, data_dir_empty: Path) -> None:
        """Ingest command should handle empty imports directory."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "ingest"],
        )

        # Should succeed with 0 files processed
        assert result.exit_code == 0
        assert "Files processed: 0" in result.output

    def test_ingest_with_xlsx_file(self, data_dir_with_xlsx: Path) -> None:
        """Ingest command should process XLSX files."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_xlsx), "ingest"],
        )

        assert result.exit_code == 0
        assert "Ingestion complete" in result.output or "[OK]" in result.output
        assert "Files processed: 1" in result.output

        # Verify transaction was created
        csv_files = list((data_dir_with_xlsx / "transactions").rglob("*.csv"))
        assert len(csv_files) == 1

    def test_ingest_from_archive_not_found(self, data_dir_empty: Path) -> None:
        """Ingest --from-archive should error when file_id not found."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "ingest", "--from-archive", "999999_1"],
        )

        assert result.exit_code == 4  # NO_DATA (file_id not in history)
        assert "not found" in result.output.lower()

    def test_ingest_archive_flag(self, data_dir_with_xlsx: Path) -> None:
        """Ingest --archive should copy files to archives directory."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_xlsx), "ingest", "--archive"],
        )

        assert result.exit_code == 0
        assert "Ingestion complete" in result.output or "[OK]" in result.output

    def test_ingest_dry_run_does_not_write_files(self, data_dir_with_xlsx: Path) -> None:
        """ingest --dry-run should preview changes without modifying the filesystem."""
        before_snapshot = _snapshot_tree(data_dir_with_xlsx)

        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_xlsx), "ingest", "--dry-run"],
        )

        assert result.exit_code == 0, result.output
        assert _snapshot_tree(data_dir_with_xlsx) == before_snapshot
        assert not list((data_dir_with_xlsx / "transactions").rglob("*.csv"))
        assert "Source XLSX files found: 1" in result.output
        assert "test_export.xlsx" in result.output
        assert "No changes written" in result.output

    def test_ingest_dry_run_json(self, data_dir_with_xlsx: Path) -> None:
        """ingest --dry-run --json should return a structured preview payload."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_xlsx), "ingest", "--dry-run", "--json"],
        )

        assert result.exit_code == 0, result.output
        payload = json.loads(result.output)
        assert payload["command"] == "ingest"
        assert payload["dry_run"] is True
        assert payload["preview"]["files_found"] == 1
        assert payload["preview"]["transactions"]["estimated_new_rows"] == 1
        assert payload["preview"]["transactions"]["estimated_dedup_skips"] == 0
        assert len(payload["preview"]["transactions"]["affected_partitions"]) == 1


class TestTagCommand:
    """Tests for the tag command."""

    def test_tag_no_rules_file(self, data_dir_empty: Path) -> None:
        """Tag command should error when rules.yaml doesn't exist."""
        # Remove the rules.yaml created by fixture
        (data_dir_empty / "rules.yaml").unlink()

        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "tag"],
        )

        assert result.exit_code == 2  # USAGE_ERROR (missing rules.yaml)
        assert "Rules file not found" in result.output or "rules.yaml" in result.output

    def test_tag_with_rules(self, data_dir_with_transactions: Path) -> None:
        """Tag command should apply rules to transactions."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "tag"],
        )

        assert result.exit_code == 0
        assert "Tagging complete" in result.output or "[OK]" in result.output
        assert "Tagged:" in result.output or "tagged" in result.output.lower()

        events = _read_audit_events(data_dir_with_transactions)
        assert len(events) == 1
        assert events[0]["event"] == "financial_mutation"
        assert events[0]["command"] == "tag"
        assert events[0]["action"] == "bulk_apply"
        assert events[0]["fields_changed"] == [
            "category_rule",
            "category_final",
            "tags_rule",
            "tags_final",
            "confidence",
            "needs_review",
        ]
        assert events[0]["change_summary"] == "bulk tag applied to transaction partitions"
        assert events[0]["changed_rows"] == 3
        assert events[0]["partition_count"] == 1
        assert events[0]["success"] is True
        rendered_events = json.dumps(events, ensure_ascii=False)
        assert "스타벅스" not in rendered_events
        assert "신한카드" not in rendered_events
        assert "5000" not in rendered_events

    def test_tag_dry_run(self, data_dir_with_transactions: Path) -> None:
        """Tag --dry-run should preview changes without modifying files."""
        # Get original CSV content
        csv_path = list((data_dir_with_transactions / "transactions").rglob("*.csv"))[0]
        original_content = csv_path.read_text()

        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "tag", "--dry-run"],
        )

        assert result.exit_code == 0
        assert "Dry-run" in result.output or "dry-run" in result.output.lower()
        assert "No changes written" in result.output or "dry-run" in result.output.lower()

        # Verify file wasn't modified
        assert csv_path.read_text() == original_content
        assert _read_audit_events(data_dir_with_transactions) == []

    def test_tag_empty_transactions(self, data_dir_empty: Path) -> None:
        """Tag command should handle empty transactions directory."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "tag"],
        )

        assert result.exit_code == 0
        assert "Total transactions: 0" in result.output


class TestTransferCommand:
    """Tests for the transfer command."""

    def test_transfer_no_transactions(self, data_dir_empty: Path) -> None:
        """Transfer command should handle empty transactions."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "transfer"],
        )

        assert result.exit_code == 0
        assert "Transfer detection complete" in result.output or "[OK]" in result.output
        assert "Transfers detected: 0" in result.output

    def test_transfer_with_data(self, data_dir_with_transactions: Path) -> None:
        """Transfer command should detect transfers in existing data."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "transfer"],
        )

        assert result.exit_code == 0
        assert "Transfer detection complete" in result.output or "[OK]" in result.output


class TestExportCommand:
    """Tests for the export command."""

    def test_export_no_data(self, data_dir_empty: Path) -> None:
        """Export command should handle empty data directory."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "export"],
        )

        # May succeed with 0 rows or fail gracefully
        # Both behaviors are acceptable for empty data
        assert result.exit_code == 0 or "error" in result.output.lower()

    def test_export_xlsx_format(self, data_dir_with_transactions: Path) -> None:
        """Export command should generate XLSX master file."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "export", "--format", "xlsx"],
        )

        assert result.exit_code == 0
        assert "Export complete" in result.output or "[OK]" in result.output

        # Verify master file was created
        exports_dir = data_dir_with_transactions / "exports"
        xlsx_files = list(exports_dir.glob("master_*.xlsx"))
        assert len(xlsx_files) == 1

    def test_export_invalid_format(self, data_dir_with_transactions: Path) -> None:
        """Export command should reject invalid format."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "export", "--format", "invalid"],
        )

        assert result.exit_code == 2  # USAGE_ERROR
        assert "Invalid format" in result.output

    def test_export_invalid_period_format(self, data_dir_with_transactions: Path) -> None:
        """Export command should reject invalid period format."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "export", "--period", "2024/10"],
        )

        assert result.exit_code == 2  # USAGE_ERROR
        assert "Invalid period format" in result.output

    def test_export_invalid_period_month(self, data_dir_with_transactions: Path) -> None:
        """Export command should reject invalid month in period."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "export", "--period", "2024-13"],
        )

        assert result.exit_code == 2  # USAGE_ERROR
        assert "Invalid month" in result.output

    def test_export_valid_period(self, data_dir_with_transactions: Path) -> None:
        """Export command should accept valid period filter."""
        result = runner.invoke(
            app,
            [
                "--data-dir",
                str(data_dir_with_transactions),
                "export",
                "--format",
                "xlsx",
                "--period",
                "2024-10",
            ],
        )

        assert result.exit_code == 0

    def test_export_dry_run_does_not_write_files(self, data_dir_with_transactions: Path) -> None:
        """export --dry-run should preview artifacts without creating files."""
        before_snapshot = _snapshot_tree(data_dir_with_transactions)

        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "export", "--dry-run"],
        )

        assert result.exit_code == 0, result.output
        assert _snapshot_tree(data_dir_with_transactions) == before_snapshot
        assert not list((data_dir_with_transactions / "exports").glob("master_*.xlsx"))
        assert not (data_dir_with_transactions / "exports" / "reports").exists()
        assert "Would generate" in result.output
        assert "master_" in result.output
        assert "No files written" in result.output

    def test_export_dry_run_json(self, data_dir_with_transactions: Path) -> None:
        """export --dry-run --json should return a structured output manifest."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_transactions), "export", "--dry-run", "--json"],
        )

        assert result.exit_code == 0, result.output
        payload = json.loads(result.output)
        assert payload["command"] == "export"
        assert payload["dry_run"] is True
        assert payload["format"] == "xlsx"
        assert payload["transaction_count"] == 3
        output_files = payload["output_files"]
        assert len(output_files) == 6
        assert output_files[0]["kind"] == "master_xlsx"
        assert output_files[0]["path"].endswith(".xlsx")
        assert any(item["path"].endswith("monthly_spend.csv") for item in output_files)


class TestRefreshCommand:
    """Tests for the refresh (full pipeline) command."""

    def test_refresh_empty_data(self, data_dir_empty: Path) -> None:
        """Refresh command should run full pipeline on empty data."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_empty), "refresh"],
        )

        assert result.exit_code == 0
        assert "파이프라인 완료" in result.output or "complete" in result.output.lower()

    def test_refresh_with_xlsx(self, data_dir_with_xlsx: Path) -> None:
        """Refresh command should process new XLSX files."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_xlsx), "refresh"],
        )

        assert result.exit_code == 0
        assert "파이프라인 완료" in result.output or "pipeline" in result.output.lower()

        # Verify outputs were created
        exports_dir = data_dir_with_xlsx / "exports"
        assert exports_dir.exists()
        xlsx_files = list(exports_dir.glob("master_*.xlsx"))
        assert len(xlsx_files) == 1

    def test_refresh_pipeline_steps_logged(self, data_dir_with_xlsx: Path) -> None:
        """Refresh command should show progress for each step."""
        result = runner.invoke(
            app,
            ["--data-dir", str(data_dir_with_xlsx), "refresh"],
        )

        assert result.exit_code == 0
        # Check that pipeline steps are mentioned
        assert "ingest" in result.output.lower() or "가져" in result.output
        assert "tag" in result.output.lower() or "태깅" in result.output

    def test_refresh_help_text(self) -> None:
        """Refresh help should describe the new primary full-pipeline command."""
        result = runner.invoke(app, ["refresh", "--help"])

        assert result.exit_code == 0
        assert "Re-process all existing data" in result.output
        assert "ingest" in result.output.lower()
        assert "--json" in result.output


class TestPipelineErrorHandling:
    """Tests for error handling in pipeline commands."""

    def test_ingest_keyboard_interrupt(self, data_dir_empty: Path) -> None:
        """Ingest should handle keyboard interrupt gracefully."""
        with patch(
            "finjuice.pipeline.ingest.pipeline.ingest_all_files",
            side_effect=KeyboardInterrupt,
        ):
            result = runner.invoke(
                app,
                ["--data-dir", str(data_dir_empty), "ingest"],
            )

            assert result.exit_code == 130
            assert "cancelled" in result.output.lower()

    def test_tag_keyboard_interrupt(self, data_dir_with_transactions: Path) -> None:
        """Tag should handle keyboard interrupt gracefully."""
        with patch(
            "finjuice.pipeline.tagging.pipeline.run_tagging",
            side_effect=KeyboardInterrupt,
        ):
            result = runner.invoke(
                app,
                ["--data-dir", str(data_dir_with_transactions), "tag"],
            )

            assert result.exit_code == 130
            assert "cancelled" in result.output.lower()

    def test_export_keyboard_interrupt(self, data_dir_with_transactions: Path) -> None:
        """Export should handle keyboard interrupt gracefully."""
        with patch(
            "finjuice.pipeline.export.master.export_master_xlsx",
            side_effect=KeyboardInterrupt,
        ):
            result = runner.invoke(
                app,
                ["--data-dir", str(data_dir_with_transactions), "export"],
            )

            assert result.exit_code == 130
            assert "cancelled" in result.output.lower()

    def test_transfer_keyboard_interrupt(self, data_dir_empty: Path) -> None:
        """Transfer should handle keyboard interrupt gracefully."""
        with patch(
            "finjuice.pipeline.transfer.detection.run_transfer_detection",
            side_effect=KeyboardInterrupt,
        ):
            result = runner.invoke(
                app,
                ["--data-dir", str(data_dir_empty), "transfer"],
            )

            assert result.exit_code == 130
            assert "cancelled" in result.output.lower()


class TestRegisterPipelineCommands:
    """Tests for register_pipeline_commands function."""

    def test_commands_registered(self) -> None:
        """All pipeline commands should be registered."""
        # Get registered commands from the app
        command_names = [cmd.name for cmd in app.registered_commands]

        assert "ingest" in command_names
        assert "tag" in command_names
        assert "transfer" in command_names
        assert "export" in command_names
        assert "refresh" in command_names
