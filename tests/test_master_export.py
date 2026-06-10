"""
Unit tests for master XLSX export.

Tests the export_master_xlsx function that generates consolidated
transaction XLSX files from CSV partitions.
"""

from pathlib import Path
from typing import Any

import polars as pl
import pytest
from openpyxl import load_workbook

from finjuice.pipeline.export.master import export_master_xlsx
from finjuice.pipeline.storage import csv_partition


@pytest.fixture
def temp_csv_base_dir(tmp_path: Path):  # type: ignore[misc]
    """Create a temporary CSV partitions directory."""
    csv_dir = tmp_path / "transactions"
    csv_dir.mkdir(parents=True, exist_ok=True)
    return csv_dir


@pytest.fixture
def sample_transactions() -> list[dict[str, Any]]:
    """Sample transaction data for testing master export.

    Note: Tags are Python lists (CSV format), not JSON strings.
    """
    return [
        {
            "row_hash": "a" * 16,
            "source_row": 1,
            "date": "2025-01-15",
            "time": "14:30",
            "type_raw": "지출",
            "type_norm": "expense",
            "major_raw": "식비",
            "minor_raw": "외식",
            "merchant_raw": "스타벅스",
            "memo_raw": "커피 구매",
            "amount": -5000,
            "account": "신한카드",
            "currency": "KRW",
            "counterparty": "스타벅스",
            "datetime": "2025-01-15T14:30:00",
            "category_rule": "카페",
            "category_final": "카페",
            "tags_rule": ["카페", "커피"],  # Python list, not JSON
            "tags_ai": ["업무식대"],
            "tags_manual": ["외식"],
            "tags_final": ["카페", "커피"],
            "confidence": 1.0,
            "needs_review": 0,
            "is_transfer": 0,
            "transfer_group_id": None,
            "file_id": "250115_1",
        },
        {
            "row_hash": "b" * 16,
            "source_row": 2,
            "date": "2025-02-20",
            "time": "19:00",
            "type_raw": "지출",
            "type_norm": "expense",
            "major_raw": "식비",
            "minor_raw": "외식",
            "merchant_raw": "맥도날드",
            "memo_raw": "저녁 식사",
            "amount": -10000,
            "account": "신한카드",
            "currency": "KRW",
            "counterparty": "맥도날드",
            "datetime": "2025-02-20T19:00:00",
            "category_rule": "외식",
            "category_final": "외식",
            "tags_rule": ["외식"],  # Python list
            "tags_ai": ["패스트푸드"],
            "tags_manual": [],
            "tags_final": ["외식"],
            "confidence": 1.0,
            "needs_review": 0,
            "is_transfer": 0,
            "transfer_group_id": None,
            "file_id": "250220_1",
        },
        # Transfer transaction
        {
            "row_hash": "c" * 16,
            "source_row": 3,
            "date": "2025-02-15",
            "time": "14:00",
            "type_raw": "이체",
            "type_norm": "transfer",
            "major_raw": "내계좌이체",
            "minor_raw": "이체",
            "merchant_raw": "신한은행",
            "memo_raw": "계좌이체",
            "amount": -100000,
            "account": "신한은행",
            "currency": "KRW",
            "counterparty": "우리은행",
            "datetime": "2025-02-15T14:00:00",
            "category_rule": None,
            "category_final": "이체",
            "tags_rule": [],  # Empty list
            "tags_ai": [],
            "tags_manual": [],
            "tags_final": [],
            "confidence": 0.0,
            "needs_review": 1,
            "is_transfer": 1,
            "transfer_group_id": "T0001",
            "file_id": "250215_1",
        },
    ]


def insert_transactions_to_csv(csv_base_dir: Path, transactions: list[dict[str, Any]]) -> None:
    """Helper function to insert test transactions into CSV partitions."""
    df = pl.DataFrame(transactions)
    csv_partition.append_transactions(csv_base_dir, df, deduplicate=False)


# Test 1: Basic export success
def test_export_master_xlsx_success(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Test basic master XLSX export with sample data."""
    # Arrange
    insert_transactions_to_csv(temp_csv_base_dir, sample_transactions)
    output_path = tmp_path / "master_20251031.xlsx"

    # Act
    row_count = export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    assert output_path.exists()
    assert row_count == 3  # 3 transactions

    # Read back and verify
    df = pl.read_excel(output_path, sheet_name="Transactions", engine="openpyxl")
    assert len(df) == 3
    assert df.row(0, named=True)["date"] == "2025-02-20"  # Most recent first


# Test 2: Storage schema contract
def test_export_master_xlsx_columns_follow_csv_schema_contract(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Master export must track the current CSV storage schema."""
    # Arrange
    insert_transactions_to_csv(temp_csv_base_dir, sample_transactions)
    output_path = tmp_path / "master.xlsx"

    # Act
    export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    df = pl.read_excel(output_path, sheet_name="Transactions", engine="openpyxl")
    assert df.columns == csv_partition.CSV_COLUMNS
    assert "source_file_path" not in df.columns


# Test 3: List to comma-separated string conversion
def test_export_master_xlsx_list_to_csv(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Test that Python tag lists are converted to comma-separated strings in XLSX."""
    # Arrange
    insert_transactions_to_csv(temp_csv_base_dir, sample_transactions)
    output_path = tmp_path / "master.xlsx"

    # Act
    export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    df = pl.read_excel(output_path, sheet_name="Transactions", engine="openpyxl")
    first_row = df.row(0, named=True)  # Most recent (Feb 20, McDonald's)

    # Check tags are comma-separated strings in XLSX
    assert first_row["tags_rule"] == "외식"
    assert first_row["tags_ai"] == "패스트푸드"
    assert first_row["tags_manual"] is None
    assert first_row["tags_final"] == "외식"

    # Check second row with multiple tags
    second_row = df.row(2, named=True)  # Oldest (Jan 15, Starbucks)
    assert second_row["tags_rule"] == "카페, 커피"
    assert second_row["tags_ai"] == "업무식대"
    assert second_row["tags_manual"] == "외식"
    assert second_row["tags_final"] == "카페, 커피"


# Test 4: Empty CSV partitions handling
def test_export_master_xlsx_empty_csv(temp_csv_base_dir: Path, tmp_path: Path) -> None:
    """Test export with empty CSV partitions.

    Note: When there are no transactions, the export function returns 0
    and does NOT create an empty XLSX file (better behavior).
    """
    # Arrange
    output_path = tmp_path / "master.xlsx"

    # Act
    row_count = export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    assert row_count == 0
    # No file is created when there are no transactions
    assert not output_path.exists()


# Test 5: Directory creation
def test_export_master_xlsx_creates_directory(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Test that parent directory is created if it doesn't exist."""
    # Arrange
    insert_transactions_to_csv(temp_csv_base_dir, sample_transactions)
    output_dir = tmp_path / "out"
    output_path = output_dir / "master.xlsx"
    assert not output_dir.exists()

    # Act
    export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    assert output_dir.exists()
    assert output_path.exists()


# Test 6: Sorting verification (date DESC, time DESC)
def test_export_master_xlsx_sorting(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Test that transactions are sorted by date and time descending."""
    # Arrange
    insert_transactions_to_csv(temp_csv_base_dir, sample_transactions)
    output_path = tmp_path / "master.xlsx"

    # Act
    export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    df = pl.read_excel(output_path, sheet_name="Transactions", engine="openpyxl")
    dates = df["date"].to_list()

    # Should be in descending order: 2025-02-20, 2025-02-15, 2025-01-15
    assert dates == ["2025-02-20", "2025-02-15", "2025-01-15"]


# Test 7: Transfer transactions included
def test_export_master_xlsx_includes_transfers(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Test that transfer transactions are included in master export."""
    # Arrange
    insert_transactions_to_csv(temp_csv_base_dir, sample_transactions)
    output_path = tmp_path / "master.xlsx"

    # Act
    export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    df = pl.read_excel(output_path, sheet_name="Transactions", engine="openpyxl")
    transfers = df.filter(pl.col("is_transfer") == 1)
    assert len(transfers) == 1
    assert transfers.row(0, named=True)["transfer_group_id"] == "T0001"


def test_export_master_xlsx_neutralizes_formula_strings_without_mutating_storage(
    temp_csv_base_dir: Path, sample_transactions: list[dict[str, Any]], tmp_path: Path
) -> None:
    """Formula-like strings are neutralized in XLSX while CSV partitions stay raw."""
    # Arrange
    dangerous_transaction = {
        **sample_transactions[0],
        "row_hash": "d" * 16,
        "date": "2025-03-01",
        "time": "12:00",
        "merchant_raw": '=HYPERLINK("https://example.invalid")',
        "memo_raw": "\t=cmd",
        "account": "+신한카드",
        "category_final": "@카페",
        "amount": -1234,
    }
    insert_transactions_to_csv(temp_csv_base_dir, [dangerous_transaction])
    output_path = tmp_path / "master.xlsx"

    # Act
    row_count = export_master_xlsx(temp_csv_base_dir, output_path)

    # Assert
    assert row_count == 1

    workbook = load_workbook(output_path, data_only=False)
    worksheet = workbook["Transactions"]
    headers = [cell.value for cell in worksheet[1]]
    exported = {
        str(header): worksheet.cell(row=2, column=index + 1) for index, header in enumerate(headers)
    }

    assert exported["merchant_raw"].value == '\'=HYPERLINK("https://example.invalid")'
    assert exported["merchant_raw"].data_type == "s"
    assert exported["memo_raw"].value == "'\t=cmd"
    assert exported["account"].value == "'+신한카드"
    assert exported["category_final"].value == "'@카페"
    assert exported["amount"].value == -1234
    assert exported["amount"].data_type == "n"

    stored_df = csv_partition.get_all_transactions(temp_csv_base_dir)
    stored_row = stored_df.row(0, named=True)
    assert stored_row["merchant_raw"] == '=HYPERLINK("https://example.invalid")'
    assert stored_row["memo_raw"] == "\t=cmd"
    assert stored_row["account"] == "+신한카드"
    assert stored_row["category_final"] == "@카페"
