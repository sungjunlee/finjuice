"""Block-level fact-builder tests for split Banksalad overview parsers."""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from finjuice.pipeline.ingest.overview.balance import _parse_balance_block
from finjuice.pipeline.ingest.overview.cashflow import _parse_cashflow_block
from finjuice.pipeline.ingest.overview.insurance import _parse_insurance_rows
from finjuice.pipeline.ingest.overview.investment import _parse_investment_rows
from finjuice.pipeline.ingest.overview.loan import _parse_loan_rows
from finjuice.pipeline.ingest.overview.models import _OverviewBlockParseContext
from finjuice.pipeline.ingest.overview.sections import (
    _build_section_facts,
    _find_numbered_sections,
)
from finjuice.pipeline.ingest.overview.snapshot import _resolve_snapshot_date
from tests.test_overview_processor import (
    _write_balance_block,
    _write_cashflow_block,
    _write_full_overview_xlsx,
    _write_overview_xlsx,
)


@contextmanager
def _open_sheet(file_path: Path) -> Iterator[Any]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        yield workbook.active
    finally:
        workbook.close()


def _block_context(file_path: Path) -> _OverviewBlockParseContext:
    return _OverviewBlockParseContext(
        sheet_name="뱅샐현황",
        snapshot_date="2026-06-15",
        file_id="260615_1",
        file_name=file_path.name,
    )


def test_balance_fact_builder_emits_typed_facts_and_projections(tmp_path: Path) -> None:
    """Balance parser builds facts when no numbered-section fact ids exist."""
    # Arrange
    file_path = tmp_path / "balance.xlsx"
    _write_overview_xlsx(file_path, include_cashflow=False)

    # Act
    with _open_sheet(file_path) as sheet:
        rows, facts, warnings = _parse_balance_block(
            sheet=sheet,
            block_context=_block_context(file_path),
            fact_ids={},
        )

    # Assert
    assert warnings == []
    assert {(row["side"], row["item_name"], row["amount"]) for row in rows} == {
        ("asset", "Synthetic Deposit", 1_250_000.0),
        ("asset", "Synthetic Fund", 450_000.0),
        ("liability", "Synthetic Loan", 300_000.0),
        ("liability", "Synthetic Card Due", 50_000.0),
    }
    assert all(row["source_fact_id"] for row in rows)
    assert {fact["block_id"] for fact in facts} == {"balance_status"}
    assert {"section_label", "cell", "table_value"} <= {fact["fact_kind"] for fact in facts}


def test_cashflow_fact_builder_emits_period_facts_and_projections(tmp_path: Path) -> None:
    """Cashflow parser builds month facts when no numbered-section fact ids exist."""
    # Arrange
    file_path = tmp_path / "cashflow.xlsx"
    _write_overview_xlsx(file_path)

    # Act
    with _open_sheet(file_path) as sheet:
        rows, facts, warnings = _parse_cashflow_block(
            sheet=sheet,
            block_context=_block_context(file_path),
            fact_ids={},
        )

    # Assert
    assert warnings == []
    assert {(row["period_month"], row["category"], row["amount"]) for row in rows} == {
        ("2026-05", "수입", 2_000_000.0),
        ("2026-05", "지출", -1_400_000.0),
        ("2026-06", "수입", 2_100_000.0),
        ("2026-06", "지출", -1_500_000.0),
    }
    assert all(row["source_fact_id"] for row in rows)
    assert {fact["block_id"] for fact in facts} == {"cashflow_monthly"}
    assert {"section_label", "cell", "table_value"} <= {fact["fact_kind"] for fact in facts}


def test_cashflow_fact_builder_keeps_facts_when_projection_is_ambiguous(tmp_path: Path) -> None:
    """Ambiguous cashflow headers still emit source facts without projections."""
    # Arrange
    file_path = tmp_path / "ambiguous_cashflow.xlsx"
    _write_overview_xlsx(file_path, ambiguous_cashflow=True)

    # Act
    with _open_sheet(file_path) as sheet:
        rows, facts, warnings = _parse_cashflow_block(
            sheet=sheet,
            block_context=_block_context(file_path),
            fact_ids={},
        )

    # Assert
    assert rows == []
    assert any("Cashflow projection skipped" in warning for warning in warnings)
    assert facts
    assert {fact["block_id"] for fact in facts} == {"cashflow_monthly"}


def test_insurance_fact_builder_links_projection_to_section_facts(tmp_path: Path) -> None:
    """Insurance rows reuse numbered-section fact ids instead of rebuilding cells."""
    # Arrange
    file_path = tmp_path / "full_overview.xlsx"
    _write_full_overview_xlsx(file_path)

    # Act
    with _open_sheet(file_path) as sheet:
        sections = _find_numbered_sections(sheet)
        section_facts = _build_section_facts(
            sheet=sheet,
            sheet_name=str(sheet.title),
            snapshot_date="2026-06-15",
            file_id="260615_1",
            sections=sections,
        )
        rows = _parse_insurance_rows(
            sheet=sheet,
            snapshot_date="2026-06-15",
            file_id="260615_1",
            sections=sections,
            fact_ids=section_facts.fact_ids,
        )

    # Assert
    assert [(row["institution"], row["policy_name"], row["paid_amount"]) for row in rows] == [
        ("Synthetic Insurer", "Synthetic Policy", 120_000.0)
    ]
    assert rows[0]["source_fact_id"] in section_facts.fact_ids.values()
    assert "insurance_status" in {fact["block_id"] for fact in section_facts.rows}


def test_investment_fact_builder_links_projection_to_section_facts(tmp_path: Path) -> None:
    """Investment rows reuse numbered-section fact ids instead of rebuilding cells."""
    # Arrange
    file_path = tmp_path / "full_overview.xlsx"
    _write_full_overview_xlsx(file_path)

    # Act
    with _open_sheet(file_path) as sheet:
        sections = _find_numbered_sections(sheet)
        section_facts = _build_section_facts(
            sheet=sheet,
            sheet_name=str(sheet.title),
            snapshot_date="2026-06-15",
            file_id="260615_1",
            sections=sections,
        )
        rows = _parse_investment_rows(
            sheet=sheet,
            snapshot_date="2026-06-15",
            file_id="260615_1",
            sections=sections,
            fact_ids=section_facts.fact_ids,
        )

    # Assert
    assert [
        (row["product_type"], row["institution"], row["product_name"], row["valuation_amount"])
        for row in rows
    ] == [("펀드", "Synthetic Securities", "Synthetic Fund A", 1_050_000.0)]
    assert rows[0]["source_fact_id"] in section_facts.fact_ids.values()
    assert "investment_status" in {fact["block_id"] for fact in section_facts.rows}


def test_loan_fact_builder_links_projection_to_section_facts(tmp_path: Path) -> None:
    """Loan rows reuse numbered-section fact ids instead of rebuilding cells."""
    # Arrange
    file_path = tmp_path / "full_overview.xlsx"
    _write_full_overview_xlsx(file_path)

    # Act
    with _open_sheet(file_path) as sheet:
        sections = _find_numbered_sections(sheet)
        section_facts = _build_section_facts(
            sheet=sheet,
            sheet_name=str(sheet.title),
            snapshot_date="2026-06-15",
            file_id="260615_1",
            sections=sections,
        )
        rows = _parse_loan_rows(
            sheet=sheet,
            snapshot_date="2026-06-15",
            file_id="260615_1",
            sections=sections,
            fact_ids=section_facts.fact_ids,
        )

    # Assert
    assert [
        (row["loan_type"], row["institution"], row["product_name"], row["balance_amount"])
        for row in rows
    ] == [("신용", "Synthetic Bank", "Synthetic Loan A", 2_500_000.0)]
    assert rows[0]["source_fact_id"] in section_facts.fact_ids.values()
    assert "loan_status" in {fact["block_id"] for fact in section_facts.rows}


def test_snapshot_date_parser_prefers_labeled_cell_over_filename(tmp_path: Path) -> None:
    """Labeled 기준일 wins over an export-range filename date."""
    # Arrange
    file_path = tmp_path / "synthetic_2025-06-07~2026-06-07.xlsx"
    import xlsxwriter

    workbook = xlsxwriter.Workbook(file_path)
    date_format = workbook.add_format({"num_format": "yyyy-mm-dd"})
    sheet = workbook.add_worksheet("뱅샐현황")
    sheet.write(0, 0, "기준일")
    sheet.write_datetime(0, 1, datetime(2026, 6, 15), date_format)
    _write_balance_block(sheet, 2)
    workbook.close()

    # Act
    with _open_sheet(file_path) as worksheet:
        resolved = _resolve_snapshot_date(
            sheet=worksheet,
            file_path=file_path,
            snapshot_date=None,
            file_mtime="2026-01-01T00:00:00",
        )

    # Assert
    assert resolved == "2026-06-15"


def test_snapshot_date_parser_uses_filename_when_sheet_has_no_label(tmp_path: Path) -> None:
    """Export filename date is used when the sheet has no labeled snapshot date."""
    # Arrange
    file_path = tmp_path / "synthetic_2025-06-07~2026-06-07.xlsx"
    import xlsxwriter

    workbook = xlsxwriter.Workbook(file_path)
    sheet = workbook.add_worksheet("뱅샐현황")
    _write_balance_block(sheet, 2)
    _write_cashflow_block(
        workbook,
        sheet,
        8,
        ambiguous_cashflow=False,
        cashflow_header_as_excel_dates=True,
    )
    workbook.close()

    # Act
    with _open_sheet(file_path) as worksheet:
        resolved = _resolve_snapshot_date(
            sheet=worksheet,
            file_path=file_path,
            snapshot_date=None,
            file_mtime="2026-01-01T00:00:00",
        )

    # Assert
    assert resolved == "2026-06-07"
