"""Banksalad overview worksheet parsing.

Public entry is ``parse_banksalad_overview``. Block parsers live under
``finjuice.pipeline.ingest.overview``; this module finds the sheet, resolves
the snapshot date, and dispatches each detected block.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import polars as pl
from openpyxl import load_workbook

from ..storage import csv_partition
from .overview.balance import _find_balance_anchor_pair, _parse_balance_block
from .overview.cashflow import _parse_cashflow_block
from .overview.cells import _iter_non_empty_cells, _normalize_cell_text
from .overview.constants import (
    _ASSET_ANCHOR,
    _LIABILITY_ANCHOR,
    _OVERVIEW_SHEET_NORMALIZED,
)
from .overview.insurance import _parse_insurance_rows
from .overview.investment import _parse_investment_rows
from .overview.loan import _parse_loan_rows
from .overview.models import _Anchor, _OverviewBlockParseContext
from .overview.sections import (
    _build_section_facts,
    _find_numbered_sections,
    _is_cashflow_anchor,
)
from .overview.snapshot import _resolve_snapshot_date
from .schemas import normalize_sheet_name


@dataclass(frozen=True)
class BanksaladOverviewParseResult:
    """Result of parsing one Banksalad overview worksheet."""

    overview_facts: pl.DataFrame
    balance: pl.DataFrame
    cashflow: pl.DataFrame
    insurance: pl.DataFrame
    investments: pl.DataFrame
    loans: pl.DataFrame
    warnings: list[str]


def parse_banksalad_overview(
    file_path: Path,
    file_id: str,
    snapshot_date: str | None = None,
    file_mtime: str | None = None,
) -> BanksaladOverviewParseResult:
    """Parse a Banksalad overview worksheet without writing storage files.

    Args:
        file_path: Source XLSX workbook path.
        file_id: Import-history file identifier to attach to parsed rows.
        snapshot_date: Optional explicit ``YYYY-MM-DD`` snapshot date.
        file_mtime: Optional source file mtime ISO timestamp used as fallback.

    Returns:
        Parsed overview facts, balance projections, cashflow projections, and
        non-fatal warnings. Missing or unrecognized overview sheets return
        typed empty DataFrames.
    """
    warnings: list[str] = []

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except (OSError, BadZipFile, ValueError) as exc:
        warnings.append(f"Failed to read overview workbook {file_path.name}: {exc}")
        return _empty_result(warnings)

    try:
        sheet = _find_overview_sheet(workbook)
        if sheet is None:
            warnings.append(f"Overview sheet not found in {file_path.name}; skipped")
            return _empty_result(warnings)

        resolved_snapshot_date = _resolve_snapshot_date(
            sheet=sheet,
            file_path=file_path,
            snapshot_date=snapshot_date,
            file_mtime=file_mtime,
        )
        section_ranges = _find_numbered_sections(sheet)
        section_facts = _build_section_facts(
            sheet=sheet,
            sheet_name=str(sheet.title),
            snapshot_date=resolved_snapshot_date,
            file_id=file_id,
            sections=section_ranges,
        )
        fact_rows: list[dict[str, Any]] = list(section_facts.rows)
        block_context = _OverviewBlockParseContext(
            sheet_name=str(sheet.title),
            snapshot_date=resolved_snapshot_date,
            file_id=file_id,
            file_name=file_path.name,
        )

        balance_rows, balance_facts, balance_warnings = _parse_balance_block(
            sheet=sheet,
            block_context=block_context,
            fact_ids=section_facts.fact_ids,
        )
        fact_rows.extend(balance_facts)
        warnings.extend(balance_warnings)

        cashflow_rows, cashflow_facts, cashflow_warnings = _parse_cashflow_block(
            sheet=sheet,
            block_context=block_context,
            fact_ids=section_facts.fact_ids,
        )
        fact_rows.extend(cashflow_facts)
        warnings.extend(cashflow_warnings)
        insurance_rows = _parse_insurance_rows(
            sheet=sheet,
            snapshot_date=resolved_snapshot_date,
            file_id=file_id,
            sections=section_ranges,
            fact_ids=section_facts.fact_ids,
        )
        investment_rows = _parse_investment_rows(
            sheet=sheet,
            snapshot_date=resolved_snapshot_date,
            file_id=file_id,
            sections=section_ranges,
            fact_ids=section_facts.fact_ids,
        )
        loan_rows = _parse_loan_rows(
            sheet=sheet,
            snapshot_date=resolved_snapshot_date,
            file_id=file_id,
            sections=section_ranges,
            fact_ids=section_facts.fact_ids,
        )

        return BanksaladOverviewParseResult(
            overview_facts=_frame_from_rows(
                fact_rows,
                csv_partition.BANKSALAD_OVERVIEW_FACT_POLARS_SCHEMA,
            ).sort(["block_id", "source_row", "source_col"]),
            balance=_frame_from_rows(
                balance_rows,
                csv_partition.BANKSALAD_BALANCE_POLARS_SCHEMA,
            ).sort(["side", "category", "item_name"]),
            cashflow=_frame_from_rows(
                cashflow_rows,
                csv_partition.BANKSALAD_CASHFLOW_POLARS_SCHEMA,
            ).sort(["period_month", "category"]),
            insurance=_frame_from_rows(
                insurance_rows,
                csv_partition.BANKSALAD_INSURANCE_POLARS_SCHEMA,
            ).sort(["institution", "policy_name"]),
            investments=_frame_from_rows(
                investment_rows,
                csv_partition.BANKSALAD_INVESTMENT_POLARS_SCHEMA,
            ).sort(["institution", "product_name"]),
            loans=_frame_from_rows(
                loan_rows,
                csv_partition.BANKSALAD_LOAN_POLARS_SCHEMA,
            ).sort(["institution", "product_name"]),
            warnings=warnings,
        )
    finally:
        workbook.close()


def _empty_result(warnings: list[str]) -> BanksaladOverviewParseResult:
    return BanksaladOverviewParseResult(
        overview_facts=pl.DataFrame(schema=csv_partition.BANKSALAD_OVERVIEW_FACT_POLARS_SCHEMA),
        balance=pl.DataFrame(schema=csv_partition.BANKSALAD_BALANCE_POLARS_SCHEMA),
        cashflow=pl.DataFrame(schema=csv_partition.BANKSALAD_CASHFLOW_POLARS_SCHEMA),
        insurance=pl.DataFrame(schema=csv_partition.BANKSALAD_INSURANCE_POLARS_SCHEMA),
        investments=pl.DataFrame(schema=csv_partition.BANKSALAD_INVESTMENT_POLARS_SCHEMA),
        loans=pl.DataFrame(schema=csv_partition.BANKSALAD_LOAN_POLARS_SCHEMA),
        warnings=warnings,
    )


def _frame_from_rows(rows: list[dict[str, Any]], schema: dict[str, Any]) -> pl.DataFrame:
    if not rows:
        return pl.DataFrame(schema=schema)
    return pl.DataFrame(rows, schema=schema)


def _find_overview_sheet(workbook: Any) -> Any | None:
    for sheet_name in workbook.sheetnames:
        if normalize_sheet_name(str(sheet_name)) == _OVERVIEW_SHEET_NORMALIZED:
            return workbook[sheet_name]

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if _has_overview_anchors(sheet):
            return sheet

    return None


def _has_overview_anchors(sheet: Any) -> bool:
    asset_anchors: list[_Anchor] = []
    liability_anchors: list[_Anchor] = []

    for row, col, value in _iter_non_empty_cells(sheet):
        normalized = _normalize_cell_text(value)
        if normalized == _ASSET_ANCHOR:
            asset_anchors.append(_Anchor(row=row, col=col, text=str(value).strip()))
        elif normalized == _LIABILITY_ANCHOR:
            liability_anchors.append(_Anchor(row=row, col=col, text=str(value).strip()))
        elif _is_cashflow_anchor(normalized):
            return True

    return _find_balance_anchor_pair(asset_anchors, liability_anchors) is not None
