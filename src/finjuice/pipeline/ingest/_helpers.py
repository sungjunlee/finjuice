"""
Internal helper utilities for the ingest pipeline.

Provides reusable helper functions used across transaction and asset processing.
"""

import logging
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

from ..constants import ASSET_DERIVED_ID_HASH_LENGTH_CHARS
from ..storage import csv_partition
from ..storage.csv_schema import POLARS_SCHEMA
from .schemas import is_asset_sheet_name

logger = logging.getLogger(__name__)


def _has_value(value: Any) -> bool:
    """Return True if value is present (not empty/null-like string)."""
    if value is None:
        return False
    text = str(value).strip()
    return text.lower() not in {"", "nan", "nat", "none", "null"}


def _normalize_id_source(value: Any) -> str:
    """Normalize source value before hashed ID derivation."""
    text = str(value).strip().lower()
    return "".join(ch for ch in text if ch not in {" ", "_", "-"})


def _build_derived_asset_id(prefix: str, source_value: Any) -> str:
    """Build deterministic derived ID from normalized source value."""
    normalized = _normalize_id_source(source_value)
    if not normalized:
        raise ValueError("Cannot derive ID from empty source value")
    digest = sha256(normalized.encode("utf-8")).hexdigest()[:ASSET_DERIVED_ID_HASH_LENGTH_CHARS]
    return f"{prefix}_{digest}"


def _parse_snapshot_date(raw_value: Any, fallback_date: str) -> str:
    """Parse snapshot_date value with fallback to file mtime date."""
    if not _has_value(raw_value):
        return fallback_date

    if hasattr(raw_value, "date"):
        return str(raw_value.date().isoformat())

    value = str(raw_value).strip()
    if "T" in value:
        value = value.split("T", maxsplit=1)[0]
    value = value.replace("/", "-")

    try:
        return datetime.fromisoformat(value).date().isoformat()
    except ValueError:
        return fallback_date


def _parse_snapshot_float(value: Any, field_name: str) -> float:
    """Parse numeric field from asset snapshot row."""
    if not _has_value(value):
        raise ValueError(f"{field_name} is missing")

    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} is not numeric: {value}") from exc


def _find_asset_sheet_name(file_path: Path) -> str | None:
    """Find first asset-like sheet name using normalization rules."""
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except (PermissionError, OSError) as exc:
        logger.warning("Failed to inspect workbook sheets (%s)", type(exc).__name__)
        return None

    try:
        for sheet_name in workbook.sheetnames:
            if is_asset_sheet_name(sheet_name):
                return str(sheet_name)
    finally:
        workbook.close()

    return None


def _empty_transactions_dataframe() -> pl.DataFrame:
    """Return an empty transaction DataFrame with the canonical schema."""
    return pl.DataFrame(schema=POLARS_SCHEMA)


def _empty_asset_snapshot_dataframe() -> pl.DataFrame:
    """Return an empty asset snapshot DataFrame with the canonical schema."""
    return pl.DataFrame(schema=csv_partition.ASSET_SNAPSHOT_POLARS_SCHEMA)


def _read_excel_sheet(file_path: Path, sheet_id: int) -> pl.DataFrame:
    """Read one Excel sheet as a DataFrame, narrowing Polars' multi-sheet stub."""
    result = pl.read_excel(file_path, sheet_id=sheet_id, engine="openpyxl", raise_if_empty=False)
    return result
