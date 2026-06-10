#!/usr/bin/env python3
"""Generate synthetic Banksalad XLSX data for E2E testing.

This script creates a realistic test dataset that:
- Matches the Banksalad export format (요약, 가계부 내역 sheets)
- Contains transactions matching sample_rules.yaml patterns
- Includes transfer pairs for pairing algorithm tests
- Spans 3 months for partition testing
- Has reproducible output (fixed seed)

Usage:
    uv run python scripts/generate_e2e_fixtures.py
    uv run python scripts/generate_e2e_fixtures.py --output custom_path.xlsx
"""

import argparse
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# Fixed seed for reproducibility
SEED = 42
random.seed(SEED)

# Output path
DEFAULT_OUTPUT = Path("tests/fixtures/e2e/synthetic_banksalad_e2e.xlsx")

# Date range: 2025-08-01 to 2025-10-31 (3 months)
START_DATE = datetime(2025, 8, 1)
END_DATE = datetime(2025, 10, 31)

# Banksalad column headers
HEADERS = ["날짜", "시간", "타입", "대분류", "소분류", "내용", "금액", "화폐", "결제수단", "메모"]


# =============================================================================
# Merchant/Transaction Templates
# =============================================================================

# Merchants that match sample_rules.yaml patterns
RULE_MATCHING_MERCHANTS = {
    "카페": [
        ("스타벅스 강남점", "식비", "카페", -5500),
        ("스타벅스 판교점", "식비", "카페", -6000),
        ("이디야커피 삼성역점", "식비", "카페", -3500),
        ("이디야 코엑스점", "식비", "카페", -4000),
    ],
    "편의점": [
        ("GS25 강남역점", "생활", "편의점", -3200),
        ("GS25 삼성점", "생활", "편의점", -2500),
        ("CU 역삼점", "생활", "편의점", -4100),
        ("세븐일레븐 테헤란로점", "생활", "편의점", -5500),
    ],
    "의료": [
        ("서울건강병원", "의료/건강", "종합병원", -45000),
        ("동네내과의원", "의료/건강", "의원", -32000),
        ("동네이비인후과", "의료/건강", "의원", -15000),
        ("우리동네약국", "의료/건강", "약국", -8500),
    ],
    "교통": [
        ("카카오택시", "교통", "택시", -12500),
        ("카카오T 블루", "교통", "택시", -18000),
        ("서울메트로", "교통", "지하철", -1400),
    ],
    "구독": [
        ("넷플릭스", "문화/여가", "OTT", -17000),
        ("스포티파이", "문화/여가", "음악", -10900),
    ],
    "공과금": [
        ("아파트관리비", "주거/통신", "관리비", -180000),
        ("한국전력공사", "주거/통신", "전기", -45000),
        ("서울도시가스", "주거/통신", "가스", -25000),
    ],
    "보험": [
        ("메트라이프생명", "금융", "보험", -150000),
        ("삼성화재", "금융", "보험", -85000),
    ],
    "외식": [
        ("맥도날드 강남점", "식비", "패스트푸드", -8900),
        ("버거킹 역삼점", "식비", "패스트푸드", -9500),
        ("본죽 삼성점", "식비", "한식", -8000),
    ],
}

# Merchants that DON'T match rules (for untagged testing)
UNTAGGED_MERCHANTS = [
    ("다이소 강남점", "생활", "생활용품", -5500),
    ("올리브영 삼성역점", "생활", "화장품", -25000),
    ("이케아 광명점", "생활", "가구", -89000),
    ("쿠팡 배송", "생활", "온라인쇼핑", -35000),
    ("마켓컬리", "생활", "온라인쇼핑", -42000),
    ("배달의민족", "식비", "배달", -23000),
    ("요기요 주문", "식비", "배달", -18500),
    ("CGV 강남점", "문화/여가", "영화", -14000),
    ("교보문고 광화문점", "문화/여가", "서적", -25000),
    ("나이키 코리아", "생활", "의류", -129000),
    ("애플스토어", "생활", "전자제품", -1500000),
    ("알라딘 중고서점", "문화/여가", "서적", -8500),
    ("무신사 스토어", "생활", "의류", -65000),
    ("당근마켓", "생활", "중고거래", -15000),
    ("네이버페이 결제", "기타", "온라인결제", -28000),
]

# Accounts/Cards
ACCOUNTS = [
    "신한카드 체크",
    "삼성카드 법인",
    "카카오뱅크 체크",
    "토스뱅크",
    "우리은행 급여통장",
]

# Income sources
INCOME_SOURCES = [
    ("월급", "수입", "급여", 5500000, "우리은행 급여통장"),
    ("이자수익", "수입", "이자", 12500, "카카오뱅크 체크"),
    ("환급금", "수입", "환급", 150000, "신한카드 체크"),
    ("부수입", "수입", "기타", 500000, "토스뱅크"),
]

# Transfer pairs (for testing transfer detection algorithm)
TRANSFER_PAIRS = [
    # (출금계좌, 입금계좌, 금액, 시간차(분))
    ("우리은행 급여통장", "카카오뱅크 체크", 1000000, 2),
    ("카카오뱅크 체크", "토스뱅크", 500000, 3),
    ("신한카드 체크", "우리은행 급여통장", 300000, 1),
    ("토스뱅크", "카카오뱅크 체크", 200000, 4),
    ("우리은행 급여통장", "삼성카드 법인", 800000, 2),
    # Edge case: 5분 경계
    ("카카오뱅크 체크", "신한카드 체크", 150000, 5),
    # Edge case: large amount
    ("우리은행 급여통장", "토스뱅크", 10000000, 1),
    # Edge case: small amount
    ("토스뱅크", "카카오뱅크 체크", 10000, 3),
]

# Number of expense rows copied verbatim so ingest exercises row_hash dedup.
DUPLICATE_ROW_COUNT = 3


# =============================================================================
# Helper Functions
# =============================================================================


def random_date() -> datetime:
    """Generate random datetime within the date range."""
    delta = END_DATE - START_DATE
    random_days = random.randint(0, delta.days)
    random_seconds = random.randint(0, 86399)  # 0-23:59:59
    return START_DATE + timedelta(days=random_days, seconds=random_seconds)


def format_date(dt: datetime) -> str:
    """Format datetime as YYYY-MM-DD."""
    return dt.strftime("%Y-%m-%d")


def format_time(dt: datetime) -> str:
    """Format datetime as HH:MM:SS."""
    return dt.strftime("%H:%M:%S")


def generate_expense_transactions() -> list[dict]:
    """Generate expense transactions matching rules."""
    transactions = []

    # Rule-matching merchants (more frequent)
    for category, merchants in RULE_MATCHING_MERCHANTS.items():
        # Each category gets 3-8 transactions
        count = random.randint(3, 8)
        for _ in range(count):
            merchant, major, minor, base_amount = random.choice(merchants)
            # Add some variance to amount (±20%)
            variance = random.uniform(0.8, 1.2)
            amount = int(base_amount * variance)
            dt = random_date()

            transactions.append(
                {
                    "날짜": format_date(dt),
                    "시간": format_time(dt),
                    "타입": "지출",
                    "대분류": major,
                    "소분류": minor,
                    "내용": merchant,
                    "금액": amount,
                    "화폐": "KRW",
                    "결제수단": random.choice(ACCOUNTS),
                    "메모": "",
                }
            )

    # Untagged merchants
    for _ in range(20):
        merchant, major, minor, base_amount = random.choice(UNTAGGED_MERCHANTS)
        variance = random.uniform(0.8, 1.2)
        amount = int(base_amount * variance)
        dt = random_date()

        transactions.append(
            {
                "날짜": format_date(dt),
                "시간": format_time(dt),
                "타입": "지출",
                "대분류": major,
                "소분류": minor,
                "내용": merchant,
                "금액": amount,
                "화폐": "KRW",
                "결제수단": random.choice(ACCOUNTS),
                "메모": "",
            }
        )

    return transactions


def generate_income_transactions() -> list[dict]:
    """Generate income transactions."""
    transactions = []

    # Monthly salary (3 months)
    for month in [8, 9, 10]:
        dt = datetime(2025, month, 25, 9, 0, 0)
        merchant, major, minor, amount, account = INCOME_SOURCES[0]  # 월급

        transactions.append(
            {
                "날짜": format_date(dt),
                "시간": format_time(dt),
                "타입": "수입",
                "대분류": major,
                "소분류": minor,
                "내용": merchant,
                "금액": amount,
                "화폐": "KRW",
                "결제수단": account,
                "메모": f"2025년 {month}월 급여",
            }
        )

    # Other income (random)
    for _ in range(5):
        merchant, major, minor, base_amount, account = random.choice(INCOME_SOURCES[1:])
        variance = random.uniform(0.8, 1.2)
        amount = int(base_amount * variance)
        dt = random_date()

        transactions.append(
            {
                "날짜": format_date(dt),
                "시간": format_time(dt),
                "타입": "수입",
                "대분류": major,
                "소분류": minor,
                "내용": merchant,
                "금액": amount,
                "화폐": "KRW",
                "결제수단": account,
                "메모": "",
            }
        )

    return transactions


def generate_transfer_transactions() -> list[dict]:
    """Generate transfer pairs for pairing algorithm testing."""
    transactions = []

    for from_account, to_account, amount, time_diff_minutes in TRANSFER_PAIRS:
        # Random base time
        base_dt = random_date()

        # Outgoing transfer (negative)
        transactions.append(
            {
                "날짜": format_date(base_dt),
                "시간": format_time(base_dt),
                "타입": "이체",
                "대분류": "내계좌이체",
                "소분류": "미분류",
                "내용": to_account,  # 이체 대상
                "금액": -amount,
                "화폐": "KRW",
                "결제수단": from_account,
                "메모": "",
            }
        )

        # Incoming transfer (positive) - after time_diff_minutes
        incoming_dt = base_dt + timedelta(minutes=time_diff_minutes)
        transactions.append(
            {
                "날짜": format_date(incoming_dt),
                "시간": format_time(incoming_dt),
                "타입": "이체",
                "대분류": "내계좌이체",
                "소분류": "미분류",
                "내용": from_account,  # 이체 출처
                "금액": amount,
                "화폐": "KRW",
                "결제수단": to_account,
                "메모": "",
            }
        )

    # Add some unpaired transfers (edge cases)
    for _ in range(3):
        dt = random_date()
        amount = random.randint(100000, 1000000)
        transactions.append(
            {
                "날짜": format_date(dt),
                "시간": format_time(dt),
                "타입": "이체",
                "대분류": "투자",
                "소분류": "미분류",
                "내용": "증권사 이체",
                "금액": -amount,
                "화폐": "KRW",
                "결제수단": random.choice(ACCOUNTS),
                "메모": "투자 이체",
            }
        )

    return transactions


def generate_duplicate_transactions(source: list[dict]) -> list[dict]:
    """Return verbatim copies of expense rows.

    Identical date/time/type/merchant/amount/currency/account fields produce an
    identical ``row_hash``, so ingest deduplicates these on the spot — this keeps
    the E2E fixture exercising the dedup path on every run.
    """
    return [dict(txn) for txn in source[:DUPLICATE_ROW_COUNT]]


def generate_malformed_transactions() -> list[dict]:
    """Return rows that each trip one ingest validation failure.

    Ingest skips malformed rows individually (logged as warnings) and continues
    the file, so these keep the E2E fixture exercising row-level resilience
    without aborting the pipeline.
    """
    base = {
        "날짜": "2025-09-15",
        "시간": "12:00:00",
        "타입": "지출",
        "대분류": "생활",
        "소분류": "기타",
        "내용": "Malformed Row Vendor",
        "금액": -10000,
        "화폐": "KRW",
        "결제수단": "신한카드 체크",
        "메모": "E2E malformed-row fixture",
    }
    return [
        {**base, "날짜": "", "내용": "Malformed: missing date"},
        {**base, "금액": "not-a-number", "내용": "Malformed: non-numeric amount"},
        {**base, "결제수단": "", "내용": "Malformed: missing account"},
    ]


def create_summary_sheet(wb: Workbook, total_expense: int, total_income: int) -> None:
    """Create the summary sheet (요약)."""
    ws = wb.create_sheet("요약", 0)

    # Title
    ws["A1"] = "가계부 요약"
    ws["A1"].font = Font(bold=True, size=14)

    # Summary data
    ws["A3"] = "기간"
    ws["B3"] = f"{format_date(START_DATE)} ~ {format_date(END_DATE)}"

    ws["A4"] = "총 지출"
    ws["B4"] = f"{abs(total_expense):,}원"

    ws["A5"] = "총 수입"
    ws["B5"] = f"{total_income:,}원"

    ws["A6"] = "순수익"
    ws["B6"] = f"{total_income + total_expense:,}원"

    ws["A8"] = "※ 이 파일은 E2E 테스트용 synthetic 데이터입니다."
    ws["A9"] = f"※ 생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A10"] = f"※ Seed: {SEED}"


def create_transactions_sheet(wb: Workbook, transactions: list[dict]) -> None:
    """Create the transactions sheet (가계부 내역)."""
    ws = wb.create_sheet("가계부 내역", 1)

    # Headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Sort by date and time
    sorted_transactions = sorted(
        transactions,
        key=lambda x: (x["날짜"], x["시간"]),
    )

    # Data rows
    for row_idx, txn in enumerate(sorted_transactions, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            ws.cell(row=row_idx, column=col_idx, value=txn[header])


def generate_xlsx(output_path: Path) -> dict:
    """Generate the synthetic XLSX file.

    Returns:
        Summary statistics dict
    """
    # Generate all transactions. Edge-case rows are appended last so the random
    # sequence behind expenses/incomes/transfers stays reproducible.
    expenses = generate_expense_transactions()
    incomes = generate_income_transactions()
    transfers = generate_transfer_transactions()
    duplicates = generate_duplicate_transactions(expenses)
    malformed = generate_malformed_transactions()

    all_transactions = expenses + incomes + transfers + duplicates + malformed

    # Calculate totals (edge-case rows are excluded — they are dropped on ingest)
    total_expense = sum(t["금액"] for t in expenses)
    total_income = sum(t["금액"] for t in incomes)

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_summary_sheet(wb, total_expense, total_income)
    create_transactions_sheet(wb, all_transactions)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save
    wb.save(output_path)

    return {
        "total_transactions": len(all_transactions),
        "expenses": len(expenses),
        "incomes": len(incomes),
        "transfers": len(transfers),
        "duplicates": len(duplicates),
        "malformed": len(malformed),
        "total_expense": total_expense,
        "total_income": total_income,
        "output_path": str(output_path),
    }


def main():
    parser = argparse.ArgumentParser(description="Generate synthetic Banksalad E2E test data")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output XLSX path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    print("Generating synthetic Banksalad data...")
    print(f"  Date range: {format_date(START_DATE)} ~ {format_date(END_DATE)}")
    print(f"  Seed: {SEED}")

    stats = generate_xlsx(args.output)

    print(f"\n✅ Generated: {stats['output_path']}")
    print(f"   Total rows: {stats['total_transactions']}")
    print(f"   - Expenses: {stats['expenses']}")
    print(f"   - Incomes: {stats['incomes']}")
    print(f"   - Transfers: {stats['transfers']}")
    print(f"   - Duplicates (deduped on ingest): {stats['duplicates']}")
    print(f"   - Malformed (skipped on ingest): {stats['malformed']}")
    print(f"   Total expense: {abs(stats['total_expense']):,}원")
    print(f"   Total income: {stats['total_income']:,}원")


if __name__ == "__main__":
    main()
